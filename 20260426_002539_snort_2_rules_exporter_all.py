#!/usr/bin/env python3
"""
FTD Snort 2 Rule Extractor (All Policies, All Rules) — single-progress UI.

Connects to an FTD device via SSH, enters expert mode, and extracts ALL
Snort 2 rules from ALL Intrusion Policies on the device — both built-in
(SID < 1,000,000) and local (SID 1,000,000~9,999,999) — into one XLSX
workbook.

UI behavior:
  Only ONE 0~100% progress gauge is shown during extraction.
  No per-policy / per-file chatter while the gauge is running.
"""

from __future__ import annotations

import base64
import binascii
import getpass
import re
import os
import shlex
import sys
import time
from datetime import datetime
from typing import Optional

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

import openpyxl
import paramiko
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

LOCAL_SID_MIN = 1_000_000
LOCAL_SID_MAX = 9_999_999

ENGINE_BASES = [
    "/ngfw/var/sf/detection_engines",
    "/var/sf/detection_engines",
    "/ngfw/var/cisco/detection_engines",
]

RULE_ACTION_RE = re.compile(r"^(alert|log|pass|drop|reject|sdrop)\s", re.IGNORECASE)
SID_RE = re.compile(r"\bsid\s*:\s*(\d+)\s*;", re.IGNORECASE)
GID_RE = re.compile(r"\bgid\s*:\s*(\d+)\s*;", re.IGNORECASE)
MSG_RE = re.compile(r'\bmsg\s*:\s*"([^"]*)"', re.IGNORECASE)

PROMPT_TOKENS = ["$ ", "# ", "> "]


def prompt_connection_info() -> tuple[str, str, str]:
    print("=" * 60)
    print("  FTD Snort 2 Rule Extractor (ALL policies, ALL rules)")
    print("=" * 60)

    ftd_ip = input("FTD IP address : ").strip()
    if not ftd_ip:
        print("[!] IP address cannot be empty.")
        sys.exit(1)

    username = input("Username       : ").strip()
    if not username:
        print("[!] Username cannot be empty.")
        sys.exit(1)

    password = getpass.getpass("Password       : ")
    if not password:
        print("[!] Password cannot be empty.")
        sys.exit(1)

    print()
    return ftd_ip, username, password


class FTDShell:
    def __init__(self, ip: str, username: str, password: str) -> None:
        self.ip = ip
        self.username = username
        self.password = password
        self._pw_quoted = shlex.quote(password)
        self._client = paramiko.SSHClient()
        self._client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self._ch: Optional[paramiko.Channel] = None

    def connect(self) -> None:
        self._client.connect(
            self.ip,
            username=self.username,
            password=self.password,
            look_for_keys=False,
            allow_agent=False,
            timeout=30,
        )
        self._ch = self._client.invoke_shell(width=65535, height=200)
        self._drain_until(PROMPT_TOKENS, timeout=15)
        self._ch.send("expert\n")
        self._drain_until(["$ ", "# "], timeout=20)
        self._ch.send("stty cols 65535\n")
        self._drain_until(["$ ", "# "], timeout=10)

    def close(self) -> None:
        self._client.close()

    def _recv_all_ready(self) -> str:
        assert self._ch is not None
        buf = ""
        while self._ch.recv_ready():
            buf += self._ch.recv(131072).decode("utf-8", errors="replace")
            time.sleep(0.05)
        return buf

    def _drain_until(self, tokens: list[str], timeout: int = 30) -> str:
        buf = ""
        deadline = time.time() + timeout
        while time.time() < deadline:
            chunk = self._recv_all_ready()
            if chunk:
                buf += chunk
                if any(t in buf for t in tokens):
                    return buf
            time.sleep(0.15)
        return buf

    def run(self, cmd: str, timeout: int = 60) -> str:
        assert self._ch is not None
        ts = int(time.time() * 1000)
        s_start = f"FTDS{ts}S"
        s_end = f"FTDE{ts}E"
        self._recv_all_ready()
        self._ch.send(f"echo {s_start}; {cmd}; echo {s_end}\n")

        buf = ""
        deadline = time.time() + timeout
        while time.time() < deadline:
            chunk = self._recv_all_ready()
            if chunk:
                buf += chunk
                if f"\r\n{s_end}" in buf:
                    break
            else:
                time.sleep(0.1)

        start_marker = f"\r\n{s_start}\r\n"
        end_marker = f"\r\n{s_end}"
        start_idx = buf.find(start_marker)
        if start_idx == -1:
            return buf
        content_start = start_idx + len(start_marker)
        end_idx = buf.find(end_marker, content_start)
        if end_idx == -1:
            return buf[content_start:]
        return buf[content_start:end_idx]

    def sudo_cmd(self, inner_cmd: str) -> str:
        return f"echo {self._pw_quoted} | sudo -S {inner_cmd} 2>/dev/null"

    def read_file_b64(
        self, filepath: str, use_sudo: bool = False, timeout: int = 300
    ) -> str:
        quoted_path = shlex.quote(filepath)
        if use_sudo:
            cmd = self.sudo_cmd(f"base64 {quoted_path}")
        else:
            cmd = f"base64 {quoted_path} 2>/dev/null"

        raw = self.run(cmd, timeout=timeout)

        b64_lines = []
        for line in raw.splitlines():
            s = line.strip()
            if s.startswith("[sudo]") or s.startswith("Password:") or not s:
                continue
            b64_lines.append(s)

        b64_data = "".join(b64_lines)
        try:
            decoded = base64.b64decode(b64_data).decode("utf-8", errors="replace")
        except (binascii.Error, ValueError):
            decoded = ""
        return decoded


def list_intrusion_policies(sh: FTDShell) -> list[tuple[str, str]]:
    policies: list[tuple[str, str]] = []
    seen_dirs: set[str] = set()
    for base in ENGINE_BASES:
        ls = sh.run(f"ls {shlex.quote(base)} 2>/dev/null", timeout=10)
        if not ls.strip():
            continue
        uuids = [u.strip() for u in ls.split() if u.strip()]
        for uuid in uuids:
            eng = f"{base}/{uuid}"
            if eng in seen_dirs:
                continue
            seen_dirs.add(eng)
            name_out = sh.run(
                f"cat {shlex.quote(eng + '/name')} 2>/dev/null", timeout=10
            ).strip()
            if not name_out:
                name_out = sh.run(
                    f"cat {shlex.quote(eng + '/policy_name')} 2>/dev/null", timeout=10
                ).strip()
            if name_out:
                policies.append((name_out, eng))
    return policies


def parse_rule_line(line: str) -> Optional[dict]:
    line = line.strip()
    if not line:
        return None
    enabled = True
    if line.startswith("#"):
        candidate = line.lstrip("#").strip()
        if not RULE_ACTION_RE.match(candidate):
            return None
        enabled = False
        line = candidate
    if not RULE_ACTION_RE.match(line):
        return None
    sid_m = SID_RE.search(line)
    if not sid_m:
        return None
    sid = int(sid_m.group(1))
    gid_m = GID_RE.search(line)
    gid = int(gid_m.group(1)) if gid_m else 1
    msg_m = MSG_RE.search(line)
    msg = msg_m.group(1) if msg_m else ""
    action_m = RULE_ACTION_RE.match(line)
    action = action_m.group(1).lower() if action_m else ""
    return {
        "gid": gid,
        "sid": sid,
        "msg": msg,
        "action": action,
        "rule": line,
        "enabled": enabled,
    }


def _list_builtin_rule_files(
    sh: FTDShell, policies: list[tuple[str, str]]
) -> list[tuple[str, str]]:
    """Return list of (policy_name, rule_file_path) for all policies."""
    out: list[tuple[str, str]] = []
    for policy_name, policy_dir in policies:
        find_out = sh.run(
            f"find {shlex.quote(policy_dir)} -name '*.rules' 2>/dev/null", timeout=20
        )
        for f in find_out.splitlines():
            f = f.strip()
            if f.endswith(".rules"):
                out.append((policy_name, f))
    return out


def _list_local_rule_files(sh: FTDShell) -> list[str]:
    """Recursive grep for local SIDs across candidate roots."""
    search_roots = ["/ngfw/var", "/var/sf", "/ngfw/var/cisco", "/etc"]
    existing_roots: list[str] = []
    for root in search_roots:
        chk = sh.run(
            sh.sudo_cmd(f"test -d {shlex.quote(root)} && echo YES || echo NO"),
            timeout=10,
        )
        if "YES" in chk:
            existing_roots.append(root)
    if not existing_roots:
        existing_roots = ["/"]

    roots_quoted = " ".join(shlex.quote(r) for r in existing_roots)
    # 7-digit SID with non-digit boundary to avoid 8+ digit SIDs.
    grep_cmd = (
        f"grep -rlE --include='*.rules' "
        f"'sid\\s*:\\s*[1-9][0-9]{{6}}[^0-9]' {roots_quoted}"
    )
    grep_out = sh.run(sh.sudo_cmd(grep_cmd), timeout=300)

    files: list[str] = []
    skip_prefixes = ("/proc/", "/sys/", "/dev/", "/run/")
    for line in grep_out.splitlines():
        p = line.strip()
        if not p or not p.endswith(".rules"):
            continue
        if p.startswith("[sudo]") or p.startswith("Password:"):
            continue
        if any(p.startswith(sp) for sp in skip_prefixes):
            continue
        files.append(p)
    return sorted(set(files))


def collect_all_rules(
    sh: FTDShell, policies: list[tuple[str, str]]
) -> list[dict]:
    """Collect built-in + local rules with a SINGLE 0~100% progress gauge."""
    rules: list[dict] = []
    seen: set[tuple[int, int]] = set()

    builtin_files = _list_builtin_rule_files(sh, policies)
    local_files = _list_local_rule_files(sh)

    total_units = len(builtin_files) + len(local_files)
    if total_units == 0:
        return rules

    with tqdm(
        total=total_units,
        desc="Extracting",
        unit="file",
        ncols=80,
        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
    ) as bar:
        for policy_name, rf in builtin_files:
            content = sh.read_file_b64(rf, use_sudo=False, timeout=300)
            for line in content.splitlines():
                r = parse_rule_line(line)
                if r is None:
                    continue
                key = (r["gid"], r["sid"])
                if key in seen:
                    continue
                seen.add(key)
                if LOCAL_SID_MIN <= r["sid"] <= LOCAL_SID_MAX:
                    r["policy"] = f"{policy_name} (local)"
                else:
                    r["policy"] = policy_name
                rules.append(r)
            bar.update(1)

        for lrf in local_files:
            content = sh.read_file_b64(lrf, use_sudo=True, timeout=300)
            for line in content.splitlines():
                r = parse_rule_line(line.strip())
                if r is None:
                    continue
                if not (LOCAL_SID_MIN <= r["sid"] <= LOCAL_SID_MAX):
                    continue
                key = (r["gid"], r["sid"])
                if key in seen:
                    for existing in rules:
                        if (existing["gid"], existing["sid"]) == key:
                            if "(local)" not in existing.get("policy", ""):
                                existing["policy"] = (
                                    f"{existing.get('policy', '')} (local)".strip()
                                )
                            break
                    continue
                seen.add(key)
                r["policy"] = "(local)"
                rules.append(r)
            bar.update(1)

    rules.sort(key=lambda r: (r["gid"], r["sid"]))
    return rules


def _thin_border() -> Border:
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)


def save_xlsx(
    rules: list[dict], ftd_ip: str, policies: list[tuple[str, str]]
) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(SCRIPT_DIR, f"{ts}_FTD_Snort2_ALL.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snort2 ALL"

    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    DATA_FONT = Font(name="Arial", size=10)
    EVEN_FILL = PatternFill("solid", start_color="DEEAF1")
    ODD_FILL = PatternFill("solid", start_color="FFFFFF")
    C_ALIGN = Alignment(horizontal="center", vertical="top")
    L_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)
    BD = _thin_border()

    COLS = [
        ("GID", 8),
        ("SID", 14),
        ("Enabled", 10),
        ("Rule State", 14),
        ("Message", 65),
        ("Rule", 155),
    ]
    for ci, (hdr, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = BD
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    ENABLED_FILL = PatternFill("solid", start_color="C6EFCE")
    DISABLED_FILL = PatternFill("solid", start_color="FFCCCC")

    # Display order: SID descending so newest local rules appear on top.
    display_rules = sorted(rules, key=lambda r: r["sid"], reverse=True)

    for ri, rule in enumerate(display_rules, 2):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        enabled_val = "Yes" if rule["enabled"] else "No"
        for ci, (val, aln) in enumerate(
            [
                (rule["gid"], C_ALIGN),
                (rule["sid"], C_ALIGN),
                (enabled_val, C_ALIGN),
                (rule.get("action", ""), C_ALIGN),
                (rule["msg"], L_ALIGN),
                (rule["rule"], L_ALIGN),
            ],
            1,
        ):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT
            c.fill = fill
            c.alignment = aln
            c.border = BD
        ws.cell(row=ri, column=3).fill = (
            ENABLED_FILL if rule["enabled"] else DISABLED_FILL
        )

    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(COLS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(display_rules) + 1}"

    ws_sum = wb.create_sheet("Summary")
    builtin_count = sum(1 for r in rules if r["sid"] < LOCAL_SID_MIN)
    local_count = sum(1 for r in rules if LOCAL_SID_MIN <= r["sid"] <= LOCAL_SID_MAX)
    enabled_count = sum(1 for r in rules if r["enabled"])
    disabled_count = len(rules) - enabled_count

    summary_rows = [
        ("Parameter", "Value"),
        ("FTD IP", ftd_ip),
        ("Policy Count", len(policies)),
        ("Policies", ", ".join(p[0] for p in policies)),
        ("Scope", "ALL (Built-in + Local, all policies)"),
        ("Snort Version", "2"),
        ("Extracted At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Rules", len(rules)),
        ("Enabled Rules", enabled_count),
        ("Disabled Rules", disabled_count),
        ("Built-in Rules (SID < 1M)", builtin_count),
        ("Local Rules (SID 1M~10M)", local_count),
    ]
    for row in summary_rows:
        ws_sum.append(row)

    for ci in range(1, 3):
        c = ws_sum.cell(row=1, column=ci)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 60
    for r in range(2, len(summary_rows) + 1):
        for ci in range(1, 3):
            ws_sum.cell(row=r, column=ci).font = Font(name="Arial", size=10)

    wb.save(output_file)
    return output_file


def main() -> None:
    ftd_ip, username, password = prompt_connection_info()

    sh = FTDShell(ftd_ip, username, password)
    try:
        sh.connect()

        policies = list_intrusion_policies(sh)
        if not policies:
            print("[!] No Intrusion Policies found on this FTD.")
            sys.exit(1)

        rules = collect_all_rules(sh, policies)

        if not rules:
            print("[!] No matching rules found.")
            sys.exit(1)

        output_file = save_xlsx(rules, ftd_ip, policies)

        builtin_count = sum(1 for r in rules if r["sid"] < LOCAL_SID_MIN)
        local_count = sum(
            1 for r in rules if LOCAL_SID_MIN <= r["sid"] <= LOCAL_SID_MAX
        )
        enabled_count = sum(1 for r in rules if r["enabled"])
        disabled_count = len(rules) - enabled_count

        print()
        print("=" * 60)
        print(f"  Output  : {output_file}")
        print(f"  Total   : {len(rules)}  "
              f"(Built-in: {builtin_count}, Local: {local_count})")
        print(f"  Enabled : {enabled_count}    Disabled: {disabled_count}")
        print("=" * 60)

    finally:
        sh.close()


if __name__ == "__main__":
    main()
